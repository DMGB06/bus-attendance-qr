"""
Genera Excel combinando lista oficial (apoderados) con BD actual (alumnos).
"""
from __future__ import annotations

import csv
import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OFFICIAL_CSV = Path(r"c:\Users\denil\Downloads\Editado alumnos.csv")
CURRENT_CSV = Path(r"c:\Users\denil\Downloads\social_bus_escolar_rows.csv")
OUTPUT_XLSX = Path(r"c:\Users\denil\Downloads\Lista_Oficial_Combinada_BusControl.xlsx")

# Correcciones manuales de alta confianza (nombre oficial -> nombre en BD)
MANUAL_ALIASES: dict[str, str] = {
    "BRIGITTE TOCTO CUEVA": "BRIGITTE TOCTON CUEVA",
    "YOHAN TOCTO CUEVA": "JOHAN TOCTON CUEVA",
    "KATSUMI SALDANA RAVELLO": "KAPSUMI SALDANA RAVELLO",
    "HEYKEL ESPIRITU ORMENO": "EYKEL ESPRITUD",
    "ADRIANO ESPIRITU ORMENO": "ADRIANO ESPIRITUD",
    "KIARA FELIX ARASCUE": "KIARA DALILA FELIX ARASCUE",
    "CRISTOFER MARTINEZ NAUPA": "CRISTHOFER MARTINEZ NAUPA",
    "GLUTILDE NAUPA DOMINGUEZ": "CLOTILDE NAUPA",
    "ZOE FLORES ATOCCSA": "ZOE FLOREZ",
    "MIA LORES ATOCCSA": "MIA FLORES ACTOSSA",
    "NARSHELL DE LA CADENA ATOCCSA": "MARSHERLL DE LA CADENA ATOCSSA",
    "DUSTIN DE LA CADENA ATOCCSA": "DUSTIN DE LA CADENA ATOCSSA",
    "ALONDRA YANAC SEGOBIA": "ALONDRA YANAC SEGOVIA",
    "DEISY DEL CARPIO IGNA": "DEYSI DEL CARPIO INGA",
    "JOSELIN UCHUYPOMA HILARIO": "HILARIA UCHUYPOMA",
    "JULIETH CARDENAS UCHUYPOMA": "YULIETH CARDENAS",
    "YOHANA CARDENAS UCHUYPOMA": "YULIETH CARDENAS",
    "KATTY PARIONA": "KATHY PARIONA",
    "KATTY JIMENEZ SUAREZ": "KATHY JIMENEZ SUAREZ",
    "MARIA MONTENEGRO GOICOCHEA": "MARIA MONTENEGRO",
    "MARIA PEZO FERNANDO": "MARIANA PEZO",
    "MARIA LLAMO MOLINA": "MARIA LLAMO MOLINA",
    "ELIANA ESPINOZA": "FRANK HUAMANI",
    "ELVIRA ORMENO": "FELICITA ORMENO",
    "EMMA QUISPE HIDALGO": "ENMA QUISPE",
    "ESMIRNA PEREZ SANDOVAL": "ESMIRNA PEVEZ",
    "YOHANA PEVES MARTINEZ": "JOHANA PEVEZ MARTINEZ",
    "AZUMI ALVITES PEVES": "AZUMI ALVITES PEVEZ",
    "SHEYLA PEVES": "SHEYLA PEVEZ",
    "STEBAN QUIROZ PEVES": "LUCAS QUIROZ OEVEZ",
    "SEBASTIAN QUIROZ PEVES": "LUCAS QUIROZ OEVEZ",
    "LUCAS QUIROZ PEVES": "LUCAS QUIROZ OEVEZ",
    "NERIO CARRASCO LEON": "NERIO CARRASCO",
    "CLARIBEL CARRASCO CHERO": "CLARIBEL CARRASC0",
    "MILAN GRANDEZ HUAMAN": "MILAN GRANDE HUAMAN",
    "BRITHNEY ROJAS HUAMAN": "BRITHANY DEL PILAR ROJAS HUAMAN",
    "FABRIZIO SILVA CHAVEZ": "FABRICIO ALEJANDRO GOMEZ CHAVEZ",
    "VANIA ESTRADA CHAVEZ": "VANIA DANIELA ESTRADA CHAVEZ",
    "CARITO QUISPE QUIROZ": "CARMEN PARIAMACHI",
    "MATIAS QUISPE QUIROZ": "CARMEN PARIAMACHI",
    "OLGA MERCADO PEREZ": "OLGA LUCIA MERCADO PEVEZ",
    "JHON ESCOBAR": "JHON ESCOBAR",
    "ALESSANDRO ESCOBAR": "ALESSANDRO ESCOBAR",
    "FERNANDO TOSCANO": "FERNANDO TOSCANO PEZO",
    "JADIEL ORE LLAMO": "JADIEL ORE LLAMO",
    "VICTOR GAMERO QUISPE": "VICTOR GAMERO QUISPE",
    "LUIS SIMON MANRIQUE": "LUIS GUSTAVO SIMON MANRIQUE",
    "ARACELY MELMA DEL CARPIO": "ARACELY MELMA DEL CARPIO",
    "DYLAN MALLA DEL CARPIO": "DILAN MALLA DEL CARPIO",
    "LUIS ORTIZ FELIX": "LUIS ORTIZ",
    "EIKEL GUEVARA GUTIERREZ": "EIKEL GUEVARA GUTIERREZ",
    "AZUMY CONTRERAS GUTIERREZ": "AZUMI CONTRERAS GUTIERREZ",
    "DIEGO RAMIREZ PARIONA": "DIEGO RAMIREZ PARIONA",
    "CELESTE CARHUAZ SAMANIEGO": "CELESTE CARHUAR SAMANIEGO",
    "ROSSY CARHUAZ SAMANIEGO": "ROSY CARHUAS",
    "ZOE PADILLA LA MADRID": "ZOE PADILLA LA MADRID",
    "CIELO CATALAN CHAVEZ": "CIELO CATALAN",
    "IVANA HUARA ESCALANTE": "IVANA HUARA ESCALANTE",
    "JOAQUIN CULLANCO CORONADO": "JOAQUIN CULLANCO",
    "LUCAS CARO PADILLA": "LUCAS CARO PADILLA",
    "GIMENA CARO PADILLA": "GIMENA CARO PADILLA",
    "PRISCILA CARO PADILLA": "PRISCILA CARO PADILLA",
    "KIT ITZEL HURTADO CCASANI": "KIT ETZEL",
    "ASTRID PEREZ QUISPE": "ASTRID PEREZ QUISPE",
    "LEY GUTIERREZ CUBA": "LEY GUTIERREZ CUBA",
    "ALEX CHUCO CHAMORRO": "ALEX CHUCO CHAMORRO",
    "OSTIN CHUCO CHAMORRO": "OSTIN CHUCO CHAMORRO",
    "EMMANUEL SING SALAS": "EMMANUEL SING SALAS",
    "EDINSON CAMPOS SALAS": "EDINSON CAMPOS SALAS",
    "MARIA JOSE QUISPE SALAS": "MARIA JOSE QUISPE SALAS",
    "LUIS SOTELO PARRA": "LUIS SOTELO PARRA",
    "CRISTHIAN CATALAN CASTRO": "CRISTHIAN CATALAN CASTRO",
    "YELITZA NARVAEZ ALLAZO": "YALITZA NARVAEZ",
    "GABRIEL NARVAEZ ALLAZO": "GIANFRANK",
    "CALEB NARVAEZ ALLAZO": "CALED NARVAEZ ALLAZO",
    "ROSA MEDINA AYLAS": "ROSA MEDINA",
    "SNAYDER MEDINA AYLAS": "SNAIDER MEDINA",
    "DERRICK TOMAYLLA VARGAS": "DERRICK TOMAYLOS VARGAS",
    "ENDERS TOMAYLLA VARGAS": "DILAN TOMAYLLA",
    "THIAGO PEVES": "THIAGO PEVES",
    "SOFIA PEVES": "SOFIA PEVES",
    "SOFIA RAMIREZ ALLAZO": "SOFIA LUANA RAMIREZ ALLAZO",
    "JOSE RAMIREZ ALLAZO": "JOSE ANTONIO RAMIREZ ALLAZO",
    "NASHLEY JULIAN CCOYLLO": "NASHLEY JULIAN CCOYLLO",
    "ANNIE QUISPE SEGURA": "ANNIE QUISPE SEGURA",
    "VICTOR QUISPE SEGURA": "VICTOR QUISPE SEGURA",
    "EDUARDO QUISPE CAMPOS": "EDUARDO QUISPE CAMPOS",
    "DOMINIK PARIMIACHI CAMPOS": "DOMINCK PARIAMACHI CAMPOS",
    "NAIM MACHACUAY CARRASCO": "NAIM MACHACUAY",
    "ORIETHA MACHACUAY CARRASCO": "ORITHA MACHACUAY CARRASCO",
    "DYLAN QUERO PINA": "DILAN QUERO PINA",
    "JARED CASIMIRO MESARES": "JARED CASIMIRO MESARES",
    "ANGELA ORMENO RAMOS": "ANGELA ORMENO RAMOS",
    "MARITA PUEDES CHIPA": "MARITA QUEDA CHIPA",
    "DAYANA CACERES MENA": "DAYANA CACERES MENA",
    "JHORDY CACERES MENA": "JHORDY CACERES MENA",
    "MILLER VIDAL BUSTOS": "MILLER VIDAL BUSTOS",
    "FERNANDO VIDAL BUSTOS": "FERNANDO VIDAL BUSTOS",
    "HENRY VIDAL BUSTOS": "HENRY SIXTO VIDAL BUSTOS",
    "ARIANA QUISPE VILLALOBOS": "ARIANA QUISPE VILLALOBOS",
    "BRYAN SORIANO VENANCIO": "JOSE SORIANO VENANCIO",
    "JOSE SORIANO VENANCIO": "JOSE SORIANO VENANCIO",
    "ROSANGELA HUAMAN HINOSTROZA": "ROSANGELA HUAMAN",
    "KENNY HUAMAN HINOSTROZA": "KENNY HUAMAN",
    "LUCAS SAMANIEGO": "LUCAS SAMANIEGO",
    "DANIELA GUTIERREZ": "DANIELA GUTIERREZ SAMANIEGO",
    "FRANK HUAMANI ESPINOZA": "FRANK HUAMANI",
    "MARIA QUISPE SALAS": "MARIA JOSE QUISPE SALAS",
    "ALIS RAMIREZ BERROCAL": "ALIS RAMIREZ BERROCAL",
    "BELEN ALLAZO CARO": "BELEN ALLAZO CARO",
    "JEREMIAS ALLAZO CARO": "JEREMIAS ALLAZO CARO",
    "LUCAS ALLAZO CARO": "LUCAS ALLAZO CARO",
    "TANNERT FERREYRA MANCO": "TANNERT FERREYRA MANCO",
    "SOFIA GUTIERREZ MANCO": "SOFIA GUTIERREZ MANCO",
    "JAMES CUCHULA FLORES": "JAMES CUCHULA",
    "SIMON CUCHULA FLORES": "JUAN CUCHULA",
    "GERARD CUEVA CABRERA": "GERARD CUEVA CABRERA",
    "DULCE LEON JIMENEZ": "DULCE LEON JIMENEZ",
    "DAYIRO MACHACUAY VILLAR": "DAYIRO MAURICIO MACHACUAY VILLAR",
    "JHERSON FELIX ARASCUE": "JHERSON FELIX ARASCUE",
    "KARINA VASQUEZ MONTENEGRO": "KARINA VASQUEZ MONTENEGRO",
    "AXEL CHAMORRO VALDES": "AXEL CHAMORRO",
}


def normalize_name(value: str) -> str:
    if not value:
        return ""
    text = value.strip().upper()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"[^A-Z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_phone(value: str) -> str:
    if not value:
        return ""
    digits = re.sub(r"\D", "", value)
    if not digits or set(digits) <= {"0", "O"}:
        return ""
    return digits[-9:] if len(digits) >= 9 else digits


def format_phone(value: str) -> str:
    digits = normalize_phone(value)
    if not digits:
        return ""
    if len(digits) == 9:
        return f"{digits[:3]} {digits[3:6]} {digits[6:]}"
    return digits


def is_placeholder(value: str | None) -> bool:
    if value is None:
        return True
    text = str(value).strip().upper()
    if not text:
        return True
    if re.fullmatch(r"[0O]+", text):
        return True
    if text in {"PENDIENTE", "N/A", "NA", "NULL", "NONE"}:
        return True
    return False


def clean_field(value: str | None) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if is_placeholder(text) else text


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def token_overlap(a: str, b: str) -> float:
    ta = set(normalize_name(a).split())
    tb = set(normalize_name(b).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / max(len(ta), len(tb))


def score_match(official_name: str, db_name: str) -> float:
    on = normalize_name(official_name)
    dn = normalize_name(db_name)
    if not on or not dn:
        return 0.0
    if on == dn:
        return 1.0
    alias = MANUAL_ALIASES.get(on)
    if alias and normalize_name(alias) == dn:
        return 0.99
    return max(similarity(on, dn), token_overlap(on, dn))


def match_status(score: float, matched: bool) -> str:
    if not matched:
        return "NUEVO - no en BD"
    if score >= 0.92:
        return "MATCH ALTO"
    if score >= 0.75:
        return "MATCH MEDIO - revisar"
    return "MATCH BAJO - revisar"


def load_official() -> list[dict]:
    rows: list[dict] = []
    with OFFICIAL_CSV.open(encoding="utf-8-sig", newline="") as f:
        for i, row in enumerate(csv.DictReader(f, delimiter=";"), start=1):
            rows.append(
                {
                    "fila_oficial": i,
                    "nombre_apoderado_oficial": row.get("nombre_apoderado", "").strip(),
                    "telefono_apoderado_oficial": format_phone(row.get("telefono_apoderado", "")),
                    "telefono_normalizado": normalize_phone(row.get("telefono_apoderado", "")),
                    "nombre_hijo_oficial": row.get("nombre_hijo", "").strip(),
                }
            )
    return rows


def load_current() -> list[dict]:
    rows: list[dict] = []
    with CURRENT_CSV.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            rows.append({k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()})
    return rows


def find_best_match(official_name: str, current_rows: list[dict], used_ids: set[str]) -> tuple[dict | None, float]:
    best_row: dict | None = None
    best_score = 0.0
    for row in current_rows:
        if row["id"] in used_ids:
            continue
        score = score_match(official_name, row["nombre_alumno"])
        if score > best_score:
            best_score = score
            best_row = row
    if best_score < 0.55:
        return None, best_score
    return best_row, best_score


def build_combined_rows() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    official = load_official()
    current = load_current()
    used_ids: set[str] = set()
    combined: list[dict] = []

    for item in official:
        match_row, score = find_best_match(item["nombre_hijo_oficial"], current, used_ids)
        if match_row:
            used_ids.add(match_row["id"])

        apoderado_final = item["nombre_apoderado_oficial"]
        telefono_final = item["telefono_apoderado_oficial"]
        if not telefono_final:
            db_phone = clean_field(match_row.get("telefono_apoderado") if match_row else "")
            telefono_final = format_phone(db_phone) if db_phone else ""

        combined.append(
            {
                "N°": item["fila_oficial"],
                "Estado match": match_status(score, bool(match_row)),
                "Confianza %": round(score * 100, 1) if match_row else 0,
                "Nombre apoderado (oficial)": apoderado_final,
                "Teléfono apoderado (oficial)": telefono_final or "SIN TELÉFONO",
                "Nombre alumno (oficial)": item["nombre_hijo_oficial"],
                "Nombre alumno (BD)": match_row["nombre_alumno"] if match_row else "",
                "Código QR": clean_field(match_row.get("codigo") if match_row else ""),
                "ID alumno (BD)": match_row["id"] if match_row else "",
                "DNI alumno": clean_field(match_row.get("dni_alumno") if match_row else ""),
                "Edad": clean_field(match_row.get("edad") if match_row else ""),
                "Sexo": clean_field(match_row.get("sexo") if match_row else ""),
                "Colegio": clean_field(match_row.get("colegio") if match_row else ""),
                "Dirección": clean_field(match_row.get("direccion") if match_row else ""),
                "Apoderado (BD anterior)": clean_field(match_row.get("nombre_apoderado") if match_row else ""),
                "Teléfono (BD anterior)": clean_field(match_row.get("telefono_apoderado") if match_row else ""),
                "Activo en BD": match_row.get("activo", "") if match_row else "",
                "Notas": clean_field(match_row.get("notas") if match_row else ""),
            }
        )

    only_db = []
    for row in current:
        if row["id"] not in used_ids:
            only_db.append(
                {
                    "Nombre alumno (BD)": row["nombre_alumno"],
                    "Código QR": clean_field(row.get("codigo")),
                    "ID alumno (BD)": row["id"],
                    "Apoderado (BD anterior)": clean_field(row.get("nombre_apoderado")),
                    "Teléfono (BD anterior)": clean_field(row.get("telefono_apoderado")),
                    "Colegio": clean_field(row.get("colegio")),
                    "Dirección": clean_field(row.get("direccion")),
                    "Activo en BD": row.get("activo", ""),
                    "Estado sugerido": "INACTIVO - no está en lista oficial",
                }
            )

    guardians: dict[str, dict] = {}
    for row in combined:
        key = normalize_phone(row["Teléfono apoderado (oficial)"]) or normalize_name(
            row["Nombre apoderado (oficial)"]
        )
        if key not in guardians:
            guardians[key] = {
                "Apoderado": row["Nombre apoderado (oficial)"],
                "Teléfono": row["Teléfono apoderado (oficial)"],
                "Cantidad hijos": 0,
                "Hijos (oficial)": [],
            }
        guardians[key]["Cantidad hijos"] += 1
        guardians[key]["Hijos (oficial)"].append(row["Nombre alumno (oficial)"])

    guardians_rows = []
    for g in sorted(guardians.values(), key=lambda x: x["Apoderado"].upper()):
        guardians_rows.append(
            {
                "Apoderado": g["Apoderado"],
                "Teléfono": g["Teléfono"],
                "Cantidad hijos": g["Cantidad hijos"],
                "Hijos": " | ".join(g["Hijos (oficial)"]),
            }
        )

    return pd.DataFrame(combined), pd.DataFrame(guardians_rows), pd.DataFrame(only_db)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    new_fill = PatternFill("solid", fgColor="FCE4D6")
    inactive_fill = PatternFill("solid", fgColor="EDEDED")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if ws.title == "Lista Combinada":
            status_col = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == "Estado match":
                    status_col = idx
                    break
            if status_col:
                for r in range(2, ws.max_row + 1):
                    status = ws.cell(r, status_col).value or ""
                    fill = None
                    if "NUEVO" in status:
                        fill = new_fill
                    elif "revisar" in status.lower():
                        fill = warn_fill
                    elif r % 2 == 0:
                        fill = alt_fill
                    if fill:
                        for c in range(1, ws.max_column + 1):
                            ws.cell(r, c).fill = fill

        if ws.title == "Solo en BD (revisar)":
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = inactive_fill

        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

    wb.save(path)


def main() -> None:
    combined_df, guardians_df, only_db_df = build_combined_rows()

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        combined_df.to_excel(writer, sheet_name="Lista Combinada", index=False)
        guardians_df.to_excel(writer, sheet_name="Apoderados e Hijos", index=False)
        only_db_df.to_excel(writer, sheet_name="Solo en BD (revisar)", index=False)

        summary = pd.DataFrame(
            [
                {"Métrica": "Alumnos en lista oficial", "Valor": len(combined_df)},
                {"Métrica": "Alumnos en BD actual", "Valor": len(only_db_df) + combined_df["ID alumno (BD)"].astype(bool).sum()},
                {"Métrica": "Matches encontrados", "Valor": int((combined_df["Estado match"] != "NUEVO - no en BD").sum())},
                {"Métrica": "Nuevos (solo en lista oficial)", "Valor": int((combined_df["Estado match"] == "NUEVO - no en BD").sum())},
                {"Métrica": "Solo en BD (candidatos inactivos)", "Valor": len(only_db_df)},
                {"Métrica": "Apoderados únicos", "Valor": len(guardians_df)},
                {"Métrica": "Sin teléfono apoderado", "Valor": int((combined_df["Teléfono apoderado (oficial)"] == "SIN TELÉFONO").sum())},
                {"Métrica": "Revisión manual sugerida", "Valor": int(combined_df["Estado match"].str.contains("revisar", case=False, na=False).sum())},
            ]
        )
        summary.to_excel(writer, sheet_name="Resumen", index=False)

    style_workbook(OUTPUT_XLSX)
    print(f"Archivo generado: {OUTPUT_XLSX}")
    print(f"Filas combinadas: {len(combined_df)}")
    print(f"Apoderados únicos: {len(guardians_df)}")
    print(f"Solo en BD: {len(only_db_df)}")


if __name__ == "__main__":
    main()
